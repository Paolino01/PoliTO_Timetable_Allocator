from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from Timetable_Allocator.DB_Connection.DbAPI import DbAPI
import pandas as pd


def export_solution_to_excel(params):
    db_api = DbAPI(params)

    df = db_api.get_joined_solution(params)

    days = ['Lun', 'Mar', 'Mer', 'Gio', 'Ven']
    time_slots = ['8.30-10.00', '10.00-11.30', '11.30-13.00', '13.00-14.30', '14.30-16.00', '16.00-17.30', '17.30-19.00']

    write_students_view(params, df, days, time_slots)

    write_teachers_view(params, df, days)

def write_students_view(params, df, days, time_slots):
    file_path = "../DB_Connection/" + params.timetable_name + '_weekly_view.xlsx'

    with pd.ExcelWriter(file_path) as writer:
        for course, dg_group in df.groupby('CorsoDiLaurea'):
            current_row = 0  # Keeps track of the current row

            sheet_name = calculate_sheet_name(course)

            for orientation, orientation_group in dg_group.groupby('Orientamento'):
                for year, year_group in orientation_group.groupby('Anno'):
                    # Create empty timetable
                    timetable = pd.DataFrame('', index=time_slots, columns=days)

                    # Fill the cells
                    for _, row in year_group.iterrows():
                        lect_type, day, slot = get_teaching_information(row, pd)

                        teaching = f"{row['titolo']} " + lect_type

                        if slot in timetable.index and day in timetable.columns:
                            existing = timetable.at[slot, day]
                            timetable.at[slot, day] = existing + ' \n ' + teaching if existing else teaching

                    # Reset index to bring 'time slots' into first column
                    timetable_reset = timetable.reset_index()
                    timetable_reset.columns.name = None

                    current_row = write_timetable_to_excel(writer, pd, timetable_reset, orientation, year, sheet_name, current_row)

    adjust_cells_appearance(file_path)

def write_teachers_view(params, df, days):
    file_path = "../DB_Connection/" + params.timetable_name + '_teachings_view.xlsx'

    with pd.ExcelWriter(file_path) as writer:
        for course, dg_group in df.groupby('CorsoDiLaurea'):
            current_row = 0  # Keeps track of the current row

            sheet_name = calculate_sheet_name(course)

            for orientation, orientation_group in dg_group.groupby('Orientamento'):
                for year, year_group in orientation_group.groupby('Anno'):
                    # Create empty timetable
                    timetable = pd.DataFrame('', index=list(set(year_group["ID_Insegnamento"])), columns=["Teaching", "Slots"])

                    for teaching, teaching_group in year_group.groupby('ID_Insegnamento'):
                        teaching_slots = []

                        teaching_group['giorno'] = pd.Categorical(teaching_group['giorno'], categories=days, ordered=True)
                        teaching_group = teaching_group.sort_values(by=['giorno', 'fasciaOraria'])

                        # Fill the cells
                        for _, row in teaching_group.iterrows():
                            lect_type, day, slot = get_teaching_information(row, pd)

                            teaching_slots.append(day + " " + slot + " " + lect_type)

                        timetable.at[teaching, "Teaching"] = teaching_group["titolo"].to_list()[0] + (" alfabetica " + teaching_group["alfabetica"].to_list()[0] if teaching_group["alfabetica"].to_list()[0] != "0" else "")
                        timetable.at[teaching, "Slots"] = " \n ".join(teaching_slots)

                    current_row = write_timetable_to_excel(writer, pd, timetable, orientation, year, sheet_name, current_row)

    adjust_cells_appearance(file_path)

def calculate_sheet_name(course):
    # Max 31 characters per Excel sheet
    sheet_name = course[:26]
    if len(course) > 26:
        sheet_name += course[-4:]

    return sheet_name

'''
    This function retrieves the day and slot for a Teaching and defines its name based on the title and the lecture type
'''
def get_teaching_information(row, pd):
    day = row['giorno']
    slot = row['fasciaOraria']

    lect_type = ""
    if row['tipo_insegnamento'] == 'L':
        lect_type = '(Lezione'
    else:
        if row['tipo_insegnamento'] == 'EA':
            lect_type = '(Esercitazione'
        else:
            if row['tipo_insegnamento'] == 'EL':
                lect_type = '(Laboratorio'

    if pd.notna(row['squadra']) and row['squadra'] != 'No squadra':
        lect_type += f" - {row['squadra']}"
    lect_type += ")"

    return lect_type, day, slot

def write_timetable_to_excel(writer, pd, timetable, orientation, year, sheet_name, current_row):
    # Write orientation and year
    pd.DataFrame([[f"Orientation: {orientation} - Year: {year}"]]).to_excel(
        writer, sheet_name=sheet_name, startrow=current_row, index=False, header=False
    )
    current_row += 1

    # Write table with headers
    timetable.to_excel(
        writer, sheet_name=sheet_name, startrow=current_row, index=False, header=True
    )
    current_row += len(timetable) + 1  # Space after the table

    # Empty rows
    current_row += 2

    return current_row

def adjust_cells_appearance(file_path):
    wb = load_workbook(file_path)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            # Set the time slots column to bold
            cell = row[0]  # First column (time slots)
            if isinstance(cell.value, str) and ('Orientation' in cell.value or '0-' in cell.value):
                cell.font = Font(bold=True)

            # Makes the rows higher, depending on the number of lines they have
            row_idx = row[0].row
            max_lines = 1
            for cell in row:
                if isinstance(cell.value, str):
                    lines = cell.value.count('\n') + 1
                    if lines > 1:
                        cell.alignment = Alignment(wrap_text=True, vertical="center")
                    else:
                        cell.alignment = Alignment(vertical="center")
                    max_lines = max(max_lines, lines)

                if isinstance(row[0].value, str) and 'Orientation' not in row[0].value:
                    cell.border = Border(top = Side(style="thin"), right = Side(style="thin"), left = Side(style="thin"), bottom = Side(style="thin"))
            ws.row_dimensions[row_idx].height = max_lines * 15

        # Makes the cells wider
        for col_idx, col_cells in enumerate(ws.columns, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 80

    wb.save(file_path)

    print("Excel file saved in " + file_path)