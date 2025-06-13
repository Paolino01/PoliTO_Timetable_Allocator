from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

from Data.DbAPI import DbAPI
import pandas as pd


def export_solution_to_excel(params):
    db_api = DbAPI(params)

    df = db_api.get_joined_solution(params)

    days = ['Lun', 'Mar', 'Mer', 'Gio', 'Ven']
    time_slots = ['8.30-10.00', '10.00-11.30', '11.30-13.00', '13.00-14.30', '14.30-16.00', '16.00-17.30', '17.30-19.00']

    with pd.ExcelWriter("../Data/" + params.timetable_name + '.xlsx') as writer:
        for course, dg_group in df.groupby('CorsoDiLaurea'):
            current_row = 0  # Keeps track of the current row
            sheet_name = course[:31]  # Max 31 characters per Excel sheet

            for orientation, orientation_group in dg_group.groupby('Orientamento'):
                for year, year_group in orientation_group.groupby('Anno'):
                    # Create empty timetable
                    timetable = pd.DataFrame('', index=time_slots, columns=days)

                    # Fill the cells
                    for _, row in year_group.iterrows():
                        day = row['giorno']
                        slot = row['fasciaOraria']
                        teaching = f"{row['titolo']}" + (" (Lezione" if row['tipo_insegnamento'] == 'L' else (" (Esercitazione" if row['tipo_insegnamento'] == 'EA' else " (Laboratorio"))
                        if pd.notna(row['squadra']) and row['squadra'] != 'No squadra':
                            teaching += f" - {row['squadra']}"
                        teaching += ")"

                        if slot in timetable.index and day in timetable.columns:
                            esistente = timetable.at[slot, day]
                            timetable.at[slot, day] = esistente + ' \n ' + teaching if esistente else teaching

                    # Reset index to bring 'time slots' into first column
                    timetable_reset = timetable.reset_index()
                    timetable_reset.columns.name = None

                    # Write orientation and year
                    pd.DataFrame([[f"Orientation: {orientation} - Year: {year}"]]).to_excel(
                        writer, sheet_name=sheet_name, startrow=current_row, index=False, header=False
                    )
                    current_row += 1

                    # Write table with headers
                    timetable_reset.to_excel(
                        writer, sheet_name=sheet_name, startrow=current_row, index=False, header=True
                    )
                    current_row += len(timetable_reset) + 1  # Spazio dopo la tabella

                    # Empty rows
                    current_row += 2

    file_path = "../Data/" + params.timetable_name + '.xlsx'
    wb = load_workbook(file_path)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            # Set the time slots column to bold
            cell = row[0]  # First column (time slots)
            if cell.row == 1:
                continue
            if isinstance(cell.value, str) and '-' in cell.value:
                cell.font = Font(bold=True)

            # Makes the rows higher, depending on the number of lines they have
            row_idx = row[0].row
            max_lines = 1
            for cell in row:
                if isinstance(cell.value, str):
                    lines = cell.value.count('\n') + 1
                    if lines > 1:
                        cell.alignment = Alignment(wrap_text=True)
                    max_lines = max(max_lines, lines)
            ws.row_dimensions[row_idx].height = max_lines * 15

        # Makes the cells wider
        for col_idx, col_cells in enumerate(ws.columns, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 80

    wb.save(file_path)